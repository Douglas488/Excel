import io
import re
from typing import Optional

from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl import load_workbook
import openpyxl

app = FastAPI(title="Excel 数据处理后端", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def clean_text(text: Optional[str]) -> str:
    if not text:
        return ""
    text = str(text)
    text = re.sub(r"[\x00-\x1f\x7f-\x9f]", " ", text)
    text = " ".join(text.split())
    return text.strip()


def clean_sku(sku: Optional[str]) -> str:
    if not sku:
        return ""
    sku = str(sku)
    sku = re.sub(r"[\x00-\x1f\x7f-\x9f]", "", sku)
    sku = sku.strip()
    sku = re.sub(r"\s+", " ", sku)
    return sku


def load_sheet_dict(sheet, key_col_letter: str, value_col_letter: str) -> dict:
    data: dict[str, str] = {}
    key_col_num = openpyxl.utils.column_index_from_string(key_col_letter)
    value_col_num = openpyxl.utils.column_index_from_string(value_col_letter)
    for row in range(1, sheet.max_row + 1):
        key_cell = sheet.cell(row=row, column=key_col_num)
        value_cell = sheet.cell(row=row, column=value_col_num)
        if key_cell.value and str(key_cell.value).strip():
            clean_key = clean_text(str(key_cell.value))
            if value_cell.value:
                clean_value = clean_sku(str(value_cell.value))
                data[clean_key] = clean_value
            else:
                data[clean_key] = value_cell.value
    return data


@app.get("/healthz")
def healthz():
    return {"status": "ok"}


@app.post("/process")
async def process_excel(
    file: UploadFile = File(...),
    sku_sheet: str = Form("Sheet1"),
    sku_title_col: str = Form("B"),
    sku_col: str = Form("C"),
    cost_sheet: str = Form("Sheet2"),
    cost_sku_col: str = Form("A"),
    cost_col: str = Form("B"),
    output_sheet: str = Form("Order details"),
    output_title_col: str = Form("A"),
    output_sku_col: str = Form("B"),
    output_cost_col: str = Form("D"),
    start_row: int = Form(2),
    end_row: int = Form(5000),
):
    try:
        contents = await file.read()
        in_mem = io.BytesIO(contents)
        wb = load_workbook(in_mem)

        # 校验工作表
        if sku_sheet not in wb.sheetnames or cost_sheet not in wb.sheetnames or output_sheet not in wb.sheetnames:
            return JSONResponse(
                status_code=400,
                content={
                    "error": "工作表不存在",
                    "sheets": wb.sheetnames,
                },
            )

        sku_ws = wb[sku_sheet]
        cost_ws = wb[cost_sheet]
        out_ws = wb[output_sheet]

        sku_data = load_sheet_dict(sku_ws, sku_title_col, sku_col)
        cost_data_sheet = load_sheet_dict(cost_ws, cost_sku_col, cost_col)

        title_col_num = openpyxl.utils.column_index_from_string(output_title_col)
        out_sku_col_num = openpyxl.utils.column_index_from_string(output_sku_col)
        out_cost_col_num = openpyxl.utils.column_index_from_string(output_cost_col)

        found_sku = 0
        found_cost = 0

        for row in range(start_row, end_row + 1):
            title_cell = out_ws.cell(row=row, column=title_col_num)
            sku_cell = out_ws.cell(row=row, column=out_sku_col_num)
            cost_cell = out_ws.cell(row=row, column=out_cost_col_num)

            if title_cell.value and str(title_cell.value).strip():
                clean_title = clean_text(str(title_cell.value))
                if clean_title in sku_data:
                    sku_cell.value = sku_data[clean_title]
                    found_sku += 1
                    sku_value = clean_sku(str(sku_cell.value))
                    if sku_value in cost_data_sheet:
                        cost_cell.value = cost_data_sheet[sku_value]
                        found_cost += 1
                    else:
                        cost_cell.value = "未找到成本"
                else:
                    sku_cell.value = "未找到SKU"
                    cost_cell.value = "未找到成本"

        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)

        filename = file.filename or "result.xlsx"
        return StreamingResponse(
            out_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=processed_{filename}",
                "X-Found-SKU": str(found_sku),
                "X-Found-Cost": str(found_cost),
            },
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
