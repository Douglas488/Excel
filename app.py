from flask import Flask, request, jsonify, send_file, make_response
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime
import tempfile
import io
import base64
from werkzeug.utils import secure_filename
import json

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# 强制手动CORS处理，确保兼容性
@app.after_request
def add_cors_headers(response):
    # 允许所有来源 - 使用通配符确保兼容性
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With, Accept, Origin'
    response.headers['Access-Control-Allow-Credentials'] = 'false'
    response.headers['Access-Control-Max-Age'] = '86400'
    return response

# 处理所有OPTIONS预检请求
@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = make_response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With, Accept, Origin'
        response.headers['Access-Control-Allow-Credentials'] = 'false'
        response.headers['Access-Control-Max-Age'] = '86400'
        return response

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def clean_text(text):
    """清理文本，去除特殊字符"""
    if not text:
        return ""
        
    text = str(text)
    
    # 去除不可见字符和换行符
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', text)
    
    # 去除多余空格
    text = ' '.join(text.split())
    
    # 去除首尾空格
    text = text.strip()
    
    return text

def clean_sku(sku):
    """清理SKU，处理特殊字符和格式问题"""
    if not sku:
        return ""
        
    sku = str(sku)
    
    # 去除不可见字符
    sku = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', sku)
    
    # 去除首尾空格
    sku = sku.strip()
    
    # 统一空格处理（将多个空格替换为单个空格）
    sku = re.sub(r'\s+', ' ', sku)
    
    return sku

def load_sheet_data(sheet, key_col, value_col):
    """加载工作表数据到字典"""
    data = {}
    key_col_num = openpyxl.utils.column_index_from_string(key_col)
    value_col_num = openpyxl.utils.column_index_from_string(value_col)
    
    for row in range(1, sheet.max_row + 1):
        key_cell = sheet.cell(row=row, column=key_col_num)
        value_cell = sheet.cell(row=row, column=value_col_num)
        
        if key_cell.value and str(key_cell.value).strip():
            # 清理键值，去除特殊字符
            clean_key = clean_text(str(key_cell.value))
            # 如果是SKU数据，也清理值
            if value_cell.value:
                clean_value = clean_sku(str(value_cell.value))
                data[clean_key] = clean_value
            else:
                data[clean_key] = value_cell.value
            
    return data

@app.route('/')
def index():
    """API 首页"""
    return jsonify({
        "message": "Excel数据处理工具 API",
        "version": "1.0.0",
        "endpoints": {
            "POST /api/upload": "上传Excel文件",
            "POST /api/process": "处理Excel数据",
            "POST /api/check-consistency": "检查数据一致性",
            "GET /api/health": "健康检查"
        }
    })

@app.route('/api/health')
def health():
    """健康检查"""
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})

@app.route('/api/check-consistency', methods=['POST'])
def check_consistency():
    """检查数据一致性"""
    try:
        data = request.get_json()
        
        if not data or 'file' not in data:
            return jsonify({"error": "缺少文件数据"}), 400
            
        # 解码文件内容
        file_content = base64.b64decode(data['file']['content'])
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name
        
        try:
            # 加载工作簿
            workbook = load_workbook(temp_file_path)
            
            # 获取配置
            sku_config = data.get('sku_config', {})
            cost_config = data.get('cost_config', {})
            
            if not sku_config or not cost_config:
                return jsonify({"error": "缺少数据源配置"}), 400
            
            # 读取SKU数据
            sku_sheet_name = sku_config.get('sheet', 'Sheet1')
            if sku_sheet_name not in workbook.sheetnames:
                return jsonify({"error": f"未找到工作表: {sku_sheet_name}"}), 400
                
            sku_sheet = workbook[sku_sheet_name]
            sku_data = []
            
            for row in sku_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 0 and row[0]:  # 确保有数据
                    sku_data.append({
                        'title': str(row[0]) if len(row) > 0 else '',
                        'sku': str(row[1]) if len(row) > 1 else ''
                    })
            
            # 读取成本数据
            cost_sheet_name = cost_config.get('sheet', 'Sheet2')
            if cost_sheet_name not in workbook.sheetnames:
                return jsonify({"error": f"未找到工作表: {cost_sheet_name}"}), 400
                
            cost_sheet = workbook[cost_sheet_name]
            cost_data = []
            
            for row in cost_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 0 and row[0]:  # 确保有数据
                    cost_data.append({
                        'sku': str(row[0]) if len(row) > 0 else '',
                        'cost': row[1] if len(row) > 1 else 0
                    })
            
            # 进行数据匹配分析
            sku_skus = set(item['sku'] for item in sku_data if item['sku'])
            cost_skus = set(item['sku'] for item in cost_data if item['sku'])
            
            matched_skus = sku_skus.intersection(cost_skus)
            unmatched_skus = sku_skus - cost_skus
            unused_cost_skus = cost_skus - sku_skus
            
            total_skus = len(sku_skus)
            matched_count = len(matched_skus)
            match_rate = f"{(matched_count / total_skus * 100):.1f}%" if total_skus > 0 else "0%"
            
            return jsonify({
                "matched_count": matched_count,
                "unmatched_count": len(unmatched_skus),
                "unused_count": len(unused_cost_skus),
                "total_skus": total_skus,
                "match_rate": match_rate,
                "matched_skus": list(matched_skus)[:10],  # 只返回前10个
                "unmatched_skus": list(unmatched_skus)[:20],  # 只返回前20个
                "unused_cost_skus": list(unused_cost_skus)[:20],  # 只返回前20个
                "message": "数据一致性检查完成"
            })
            
        finally:
            # 清理临时文件
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)
                
    except Exception as e:
        return jsonify({"error": f"数据一致性检查失败: {str(e)}"}), 500

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传Excel文件并返回工作表信息"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "没有文件被上传"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "没有选择文件"}), 400
        
        if not allowed_file(file.filename):
            return jsonify({"error": "不支持的文件格式，请上传 .xlsx 或 .xls 文件"}), 400
        
        # 保存临时文件
        filename = secure_filename(file.filename)
        temp_path = os.path.join(tempfile.gettempdir(), filename)
        file.save(temp_path)
        
        # 加载工作簿
        workbook = load_workbook(temp_path)
        sheet_names = workbook.sheetnames
        
        # 获取每个工作表的基本信息
        sheets_info = {}
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            sheets_info[sheet_name] = {
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "columns": [openpyxl.utils.get_column_letter(col) for col in range(1, sheet.max_column + 1)]
            }
        
        # 清理临时文件
        os.remove(temp_path)
        
        return jsonify({
            "message": "文件上传成功",
            "filename": filename,
            "sheets": sheet_names,
            "sheets_info": sheets_info
        })
        
    except Exception as e:
        return jsonify({"error": f"文件上传失败: {str(e)}"}), 500

@app.route('/api/process', methods=['POST'])
def process_data():
    """处理Excel数据"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "请求数据为空"}), 400
        
        # 验证必需参数
        required_fields = ['file', 'sku_config', 'cost_config', 'output_config']
        for field in required_fields:
            if field not in data:
                return jsonify({"error": f"缺少必需参数: {field}"}), 400
        
        # 处理文件数据
        file_data = data['file']
        if not file_data or 'content' not in file_data:
            return jsonify({"error": "文件数据无效"}), 400
        
        # 从base64解码文件内容
        import base64
        try:
            file_content = base64.b64decode(file_data['content'])
        except Exception as e:
            return jsonify({"error": f"Base64解码失败: {str(e)}"}), 400
        
        # 验证文件内容
        if len(file_content) == 0:
            return jsonify({"error": "文件内容为空"}), 400
        
        # 检查文件头是否为Excel格式
        if not (file_content.startswith(b'PK') or file_content.startswith(b'\xd0\xcf\x11\xe0')):
            return jsonify({"error": "文件格式不正确，请确保上传的是Excel文件"}), 400
        
        # 创建临时文件
        temp_path = os.path.join(tempfile.gettempdir(), f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        try:
            with open(temp_path, 'wb') as f:
                f.write(file_content)
        except Exception as e:
            return jsonify({"error": f"创建临时文件失败: {str(e)}"}), 500
        
        # 加载工作簿
        try:
            workbook = load_workbook(temp_path)
        except Exception as e:
            return jsonify({"error": f"无法打开Excel文件: {str(e)}"}), 400
        
        # 获取配置
        sku_config = data['sku_config']
        cost_config = data['cost_config']
        output_config = data['output_config']
        
        # 加载SKU数据
        sku_sheet = workbook[sku_config['sheet']]
        sku_data = load_sheet_data(sku_sheet, sku_config['title_col'], sku_config['sku_col'])
        
        # 加载成本数据
        cost_sheet = workbook[cost_config['sheet']]
        cost_data = load_sheet_data(cost_sheet, cost_config['sku_col'], cost_config['cost_col'])
        
        # 处理输出工作表
        output_sheet = workbook[output_config['sheet']]
        start_row = output_config.get('start_row', 2)
        end_row = output_config.get('end_row', 5000)
        
        # 获取列号
        title_col_num = openpyxl.utils.column_index_from_string(output_config['title_col'])
        sku_col_num = openpyxl.utils.column_index_from_string(output_config['sku_col'])
        cost_col_num = openpyxl.utils.column_index_from_string(output_config['cost_col'])
        
        processed = 0
        found_sku_count = 0
        found_cost_count = 0
        
        # 处理每一行
        for row in range(start_row, end_row + 1):
            title_cell = output_sheet.cell(row=row, column=title_col_num)
            sku_cell = output_sheet.cell(row=row, column=sku_col_num)
            cost_cell = output_sheet.cell(row=row, column=cost_col_num)
            
            if title_cell.value and str(title_cell.value).strip():
                # 清理标题文本
                clean_title = clean_text(str(title_cell.value))
                
                # 查找SKU
                if clean_title in sku_data:
                    sku_cell.value = sku_data[clean_title]
                    found_sku_count += 1
                    
                    # 根据SKU查找成本
                    sku_value = clean_sku(str(sku_cell.value))
                    if sku_value in cost_data:
                        cost_cell.value = cost_data[sku_value]
                        found_cost_count += 1
                    else:
                        cost_cell.value = "未找到成本"
                else:
                    sku_cell.value = "未找到SKU"
                    cost_cell.value = "未找到成本"
            
            processed += 1
        
        # 保存处理后的文件
        output_filename = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), output_filename)
        workbook.save(output_path)
        
        # 读取处理后的文件内容
        with open(output_path, 'rb') as f:
            output_content = f.read()
        
        # 清理临时文件
        os.remove(temp_path)
        os.remove(output_path)
        
        return jsonify({
            "message": "数据处理完成",
            "processed_rows": processed,
            "found_sku": found_sku_count,
            "found_cost": found_cost_count,
            "output_file": {
                "filename": output_filename,
                "content": base64.b64encode(output_content).decode('utf-8')
            }
        })
        
    except Exception as e:
        return jsonify({"error": f"数据处理失败: {str(e)}"}), 500

@app.route('/api/download/<filename>')
def download_file(filename):
    """下载处理后的文件"""
    try:
        # 这里应该从安全的存储位置获取文件
        # 为了演示，我们返回一个错误
        return jsonify({"error": "文件下载功能需要实现安全存储"}), 501
    except Exception as e:
        return jsonify({"error": f"文件下载失败: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
