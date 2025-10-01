import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime

class ExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel数据处理工具 - SKU查找与成本获取")
        self.root.geometry("1500x800")
        self.root.configure(bg='#2c3e50')
        
        # 设置现代化字体
        self.font_title = ("微软雅黑", 18, "bold")
        self.font_subtitle = ("微软雅黑", 12, "bold")
        self.font_normal = ("微软雅黑", 10)
        self.font_small = ("微软雅黑", 9)
        
        # 数据存储
        self.workbook = None
        self.sheet_names = []
        self.sku_data = {}
        self.cost_data = {}
        
        self.setup_ui()
        
    def setup_ui(self):
        # 主标题区域
        header_frame = tk.Frame(self.root, bg='#34495e', height=100)
        header_frame.pack(fill='x', padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="📊 Excel数据处理工具", 
                              font=self.font_title, 
                              bg='#34495e', fg='#ecf0f1')
        title_label.pack(pady=(15, 5))
        
        subtitle_label = tk.Label(header_frame, text="SKU查找与成本获取 - 现代化数据处理解决方案", 
                                 font=self.font_normal, 
                                 bg='#34495e', fg='#bdc3c7')
        subtitle_label.pack(pady=(0, 15))
        
        # 主内容区域
        main_frame = tk.Frame(self.root, bg='#2c3e50')
        main_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # 文件选择卡片
        file_card = self.create_card(main_frame, "📁 文件选择")
        
        file_select_frame = tk.Frame(file_card, bg='#ecf0f1')
        file_select_frame.pack(fill='x', padx=15, pady=10)
        
        self.file_path_var = tk.StringVar()
        self.file_entry = tk.Entry(file_select_frame, textvariable=self.file_path_var, 
                                  font=self.font_normal, width=60, 
                                  relief='solid', bd=1, bg='#ffffff', fg='#2c3e50')
        self.file_entry.pack(side='left', padx=(0, 10), pady=5)
        
        self.browse_btn = tk.Button(file_select_frame, text="📁 浏览文件", 
                                   command=self.browse_file,
                                   font=self.font_normal, bg='#3498db', fg='#ffffff',
                                   relief='flat', padx=20, pady=8, cursor='hand2',
                                   activebackground='#2980b9', activeforeground='#ffffff')
        self.browse_btn.pack(side='left')
        
        # 数据源配置区域 - 横向排列
        config_frame = tk.Frame(main_frame, bg='#2c3e50')
        config_frame.pack(fill='x', pady=10)
        
        # SKU数据源卡片
        sku_card = self.create_card(config_frame, "🔍 SKU数据源")
        self.setup_sku_config(sku_card)
        
        # 成本数据源卡片
        cost_card = self.create_card(config_frame, "💰 成本数据源")
        self.setup_cost_config(cost_card)
        
        # 输出设置卡片
        output_card = self.create_card(config_frame, "📤 输出设置")
        self.setup_output_config(output_card)
        
        # 操作控制区域
        control_frame = tk.Frame(main_frame, bg='#2c3e50')
        control_frame.pack(fill='x', pady=20)
        
        self.setup_control_buttons(control_frame)
        
        # 状态显示区域
        status_frame = tk.Frame(main_frame, bg='#2c3e50')
        status_frame.pack(fill='x', pady=10)
        
        self.setup_status_area(status_frame)
        
        # 绑定事件
        self.sku_sheet_combo.bind('<<ComboboxSelected>>', self.on_sheet_selected)
        self.cost_sheet_combo.bind('<<ComboboxSelected>>', self.on_sheet_selected)
        self.output_sheet_combo.bind('<<ComboboxSelected>>', self.on_output_sheet_selected)
        
    def create_card(self, parent, title):
        """创建现代化卡片"""
        card = tk.Frame(parent, bg='#ecf0f1', relief='solid', bd=1)
        card.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        
        # 卡片标题
        title_label = tk.Label(card, text=title, 
                              font=self.font_subtitle, 
                              bg='#ecf0f1', fg='#2c3e50')
        title_label.pack(pady=(15, 10))
        
        return card
        
    def setup_sku_config(self, parent):
        """设置SKU数据源配置"""
        # 工作表选择
        sheet_frame = tk.Frame(parent, bg='#ecf0f1')
        sheet_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(sheet_frame, text="工作表:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.sku_sheet_var = tk.StringVar()
        self.sku_sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sku_sheet_var, 
                                           state='readonly', width=12, font=self.font_normal)
        self.sku_sheet_combo.pack(side='left', padx=(5, 15))
        
        # 标题列选择
        title_frame = tk.Frame(parent, bg='#ecf0f1')
        title_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(title_frame, text="标题列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.sku_title_col_var = tk.StringVar()
        self.sku_title_col_combo = ttk.Combobox(title_frame, textvariable=self.sku_title_col_var, 
                                               state='readonly', width=8, font=self.font_normal)
        self.sku_title_col_combo.pack(side='left', padx=(5, 15))
        
        # SKU列选择
        sku_frame = tk.Frame(parent, bg='#ecf0f1')
        sku_frame.pack(fill='x', padx=15, pady=(5, 15))
        
        tk.Label(sku_frame, text="SKU列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.sku_col_var = tk.StringVar()
        self.sku_col_combo = ttk.Combobox(sku_frame, textvariable=self.sku_col_var, 
                                         state='readonly', width=8, font=self.font_normal)
        self.sku_col_combo.pack(side='left', padx=5)
        
    def setup_cost_config(self, parent):
        """设置成本数据源配置"""
        # 工作表选择
        sheet_frame = tk.Frame(parent, bg='#ecf0f1')
        sheet_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(sheet_frame, text="工作表:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.cost_sheet_var = tk.StringVar()
        self.cost_sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.cost_sheet_var, 
                                            state='readonly', width=12, font=self.font_normal)
        self.cost_sheet_combo.pack(side='left', padx=(5, 15))
        
        # SKU列选择
        sku_frame = tk.Frame(parent, bg='#ecf0f1')
        sku_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(sku_frame, text="SKU列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.cost_sku_col_var = tk.StringVar()
        self.cost_sku_col_combo = ttk.Combobox(sku_frame, textvariable=self.cost_sku_col_var, 
                                              state='readonly', width=8, font=self.font_normal)
        self.cost_sku_col_combo.pack(side='left', padx=(5, 15))
        
        # 成本列选择
        cost_frame = tk.Frame(parent, bg='#ecf0f1')
        cost_frame.pack(fill='x', padx=15, pady=(5, 15))
        
        tk.Label(cost_frame, text="成本列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.cost_col_var = tk.StringVar()
        self.cost_col_combo = ttk.Combobox(cost_frame, textvariable=self.cost_col_var, 
                                          state='readonly', width=8, font=self.font_normal)
        self.cost_col_combo.pack(side='left', padx=5)
        
    def setup_output_config(self, parent):
        """设置输出配置"""
        # 输出工作表
        sheet_frame = tk.Frame(parent, bg='#ecf0f1')
        sheet_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(sheet_frame, text="输出工作表:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.output_sheet_var = tk.StringVar()
        self.output_sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.output_sheet_var, 
                                              state='readonly', width=12, font=self.font_normal)
        self.output_sheet_combo.pack(side='left', padx=(5, 15))
        
        # 标题列
        title_frame = tk.Frame(parent, bg='#ecf0f1')
        title_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(title_frame, text="标题列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.output_title_col_var = tk.StringVar()
        self.output_title_col_combo = ttk.Combobox(title_frame, textvariable=self.output_title_col_var, 
                                                  state='readonly', width=8, font=self.font_normal)
        self.output_title_col_combo.pack(side='left', padx=(5, 15))
        
        # SKU列
        sku_frame = tk.Frame(parent, bg='#ecf0f1')
        sku_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(sku_frame, text="SKU列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.output_sku_col_var = tk.StringVar()
        self.output_sku_col_combo = ttk.Combobox(sku_frame, textvariable=self.output_sku_col_var, 
                                                state='readonly', width=8, font=self.font_normal)
        self.output_sku_col_combo.pack(side='left', padx=(5, 15))
        
        # 成本列
        cost_frame = tk.Frame(parent, bg='#ecf0f1')
        cost_frame.pack(fill='x', padx=15, pady=5)
        
        tk.Label(cost_frame, text="成本列:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.output_cost_col_var = tk.StringVar()
        self.output_cost_col_combo = ttk.Combobox(cost_frame, textvariable=self.output_cost_col_var, 
                                                 state='readonly', width=8, font=self.font_normal)
        self.output_cost_col_combo.pack(side='left', padx=(5, 15))
        
        # 行范围设置
        range_frame = tk.Frame(parent, bg='#ecf0f1')
        range_frame.pack(fill='x', padx=15, pady=(5, 15))
        
        tk.Label(range_frame, text="起始行:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.start_row_var = tk.StringVar(value="2")
        self.start_row_entry = tk.Entry(range_frame, textvariable=self.start_row_var, 
                                       width=8, font=self.font_normal, relief='solid', bd=1)
        self.start_row_entry.pack(side='left', padx=(5, 15))
        
        tk.Label(range_frame, text="结束行:", 
                font=self.font_normal, bg='#ecf0f1', fg='#2c3e50').pack(side='left')
        
        self.end_row_var = tk.StringVar(value="5000")
        self.end_row_entry = tk.Entry(range_frame, textvariable=self.end_row_var, 
                                     width=8, font=self.font_normal, relief='solid', bd=1)
        self.end_row_entry.pack(side='left', padx=5)
        
    def setup_control_buttons(self, parent):
        """设置操作控制按钮"""
        # 主要操作按钮
        main_buttons_frame = tk.Frame(parent, bg='#2c3e50')
        main_buttons_frame.pack(pady=10)
        
        self.auto_config_btn = tk.Button(main_buttons_frame, text="🎯 自动配置", 
                                        command=self.auto_config,
                                        font=self.font_normal, bg='#16a085', fg='#ffffff',
                                        relief='flat', padx=20, pady=10, cursor='hand2',
                                        activebackground='#138d75', activeforeground='#ffffff')
        self.auto_config_btn.pack(side='left', padx=5)
        
        self.check_data_btn = tk.Button(main_buttons_frame, text="🔍 数据检查", 
                                       command=self.check_data_consistency,
                                       font=self.font_normal, bg='#e67e22', fg='#ffffff',
                                       relief='flat', padx=20, pady=10, cursor='hand2',
                                       activebackground='#d35400', activeforeground='#ffffff')
        self.check_data_btn.pack(side='left', padx=5)
        
        self.load_btn = tk.Button(main_buttons_frame, text="📥 加载数据", 
                                 command=self.load_data,
                                 font=self.font_normal, bg='#27ae60', fg='#ffffff',
                                 relief='flat', padx=20, pady=10, cursor='hand2',
                                 activebackground='#229954', activeforeground='#ffffff')
        self.load_btn.pack(side='left', padx=5)
        
        self.process_btn = tk.Button(main_buttons_frame, text="⚡ 开始处理", 
                                    command=self.process_data,
                                    font=self.font_normal, bg='#e74c3c', fg='#ffffff',
                                    relief='flat', padx=20, pady=10, cursor='hand2',
                                    activebackground='#c0392b', activeforeground='#ffffff')
        self.process_btn.pack(side='left', padx=5)
        
        self.save_btn = tk.Button(main_buttons_frame, text="💾 保存结果", 
                                 command=self.save_results,
                                 font=self.font_normal, bg='#f39c12', fg='#ffffff',
                                 relief='flat', padx=20, pady=10, cursor='hand2',
                                 activebackground='#e67e22', activeforeground='#ffffff')
        self.save_btn.pack(side='left', padx=5)
        
        # 工具按钮
        tool_buttons_frame = tk.Frame(parent, bg='#2c3e50')
        tool_buttons_frame.pack(pady=5)
        
        self.debug_btn = tk.Button(tool_buttons_frame, text="🔧 调试信息", 
                                  command=self.show_debug_info,
                                  font=self.font_small, bg='#9b59b6', fg='#ffffff',
                                  relief='flat', padx=15, pady=8, cursor='hand2',
                                  activebackground='#8e44ad', activeforeground='#ffffff')
        self.debug_btn.pack(side='left', padx=5)
        
        self.test_save_btn = tk.Button(tool_buttons_frame, text="🧪 测试保存", 
                                      command=self.test_save,
                                      font=self.font_small, bg='#34495e', fg='#ffffff',
                                      relief='flat', padx=15, pady=8, cursor='hand2',
                                      activebackground='#2c3e50', activeforeground='#ffffff')
        self.test_save_btn.pack(side='left', padx=5)
        
        self.help_btn = tk.Button(tool_buttons_frame, text="📖 使用说明", 
                                 command=self.show_help,
                                 font=self.font_small, bg='#e74c3c', fg='#ffffff',
                                 relief='flat', padx=15, pady=8, cursor='hand2',
                                 activebackground='#c0392b', activeforeground='#ffffff')
        self.help_btn.pack(side='left', padx=5)
        
    def setup_status_area(self, parent):
        """设置状态显示区域"""
        # 创建状态卡片
        status_card = tk.Frame(parent, bg='#ecf0f1', relief='solid', bd=1)
        status_card.pack(fill='x', padx=10, pady=5)
        
        # 卡片标题
        title_label = tk.Label(status_card, text="📊 处理状态", 
                              font=self.font_subtitle, 
                              bg='#ecf0f1', fg='#2c3e50')
        title_label.pack(pady=(15, 10))
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(status_card, variable=self.progress_var, 
                                           maximum=100, length=400)
        self.progress_bar.pack(pady=10)
        
        # 状态标签
        self.status_var = tk.StringVar(value="请选择Excel文件开始处理")
        self.status_label = tk.Label(status_card, textvariable=self.status_var, 
                                    font=self.font_normal, bg='#ecf0f1', fg='#7f8c8d')
        self.status_label.pack(pady=5)
        
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.load_workbook()
            
    def load_workbook(self):
        try:
            self.workbook = load_workbook(self.file_path_var.get())
            self.sheet_names = self.workbook.sheetnames
            
            # 更新工作表下拉框
            self.sku_sheet_combo['values'] = self.sheet_names
            self.cost_sheet_combo['values'] = self.sheet_names
            self.output_sheet_combo['values'] = self.sheet_names
            
            if self.sheet_names:
                # 设置默认工作表
                if 'Sheet1' in self.sheet_names:
                    self.sku_sheet_combo.set('Sheet1')
                else:
                    self.sku_sheet_combo.set(self.sheet_names[0])
                    
                if 'Sheet2' in self.sheet_names:
                    self.cost_sheet_combo.set('Sheet2')
                else:
                    self.cost_sheet_combo.set(self.sheet_names[0])
                    
                if 'Order details' in self.sheet_names:
                    self.output_sheet_combo.set('Order details')
                else:
                    self.output_sheet_combo.set(self.sheet_names[0])
                
            self.status_var.set(f"已加载工作簿，共{len(self.sheet_names)}个工作表")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载工作簿失败: {str(e)}")
            
    def on_sheet_selected(self, event=None):
        # 当选择工作表时，更新列选项
        self.update_column_options()
        
    def on_output_sheet_selected(self, event=None):
        # 当选择输出工作表时，更新输出列选项
        self.update_output_column_options()
        
    def update_column_options(self):
        try:
            if not self.workbook:
                return
                
            # 更新SKU数据源的列选项
            if self.sku_sheet_var.get():
                sku_sheet = self.workbook[self.sku_sheet_var.get()]
                sku_columns = self.get_column_letters(sku_sheet)
                self.sku_title_col_combo['values'] = sku_columns
                self.sku_col_combo['values'] = sku_columns
                
            # 更新成本数据源的列选项
            if self.cost_sheet_var.get():
                cost_sheet = self.workbook[self.cost_sheet_var.get()]
                cost_columns = self.get_column_letters(cost_sheet)
                self.cost_sku_col_combo['values'] = cost_columns
                self.cost_col_combo['values'] = cost_columns
                
        except Exception as e:
            print(f"更新列选项时出错: {e}")
            
    def update_output_column_options(self):
        """更新输出工作表的列选项"""
        try:
            if not self.workbook:
                return
                
            # 更新输出工作表的列选项
            if self.output_sheet_var.get():
                output_sheet = self.workbook[self.output_sheet_var.get()]
                output_columns = self.get_column_letters(output_sheet)
                self.output_title_col_combo['values'] = output_columns
                self.output_sku_col_combo['values'] = output_columns
                self.output_cost_col_combo['values'] = output_columns
                
        except Exception as e:
            print(f"更新输出列选项时出错: {e}")
            
    def get_column_letters(self, sheet):
        """获取工作表的列字母"""
        columns = []
        for col in range(1, sheet.max_column + 1):
            columns.append(openpyxl.utils.get_column_letter(col))
        return columns
        
    def load_data(self):
        """加载数据到内存"""
        try:
            if not self.workbook:
                messagebox.showwarning("警告", "请先选择Excel文件")
                return
                
            self.status_var.set("正在加载数据...")
            self.progress_var.set(0)
            self.root.update()
            
            # 加载SKU数据
            if self.sku_sheet_var.get() and self.sku_title_col_var.get() and self.sku_col_var.get():
                sku_sheet = self.workbook[self.sku_sheet_var.get()]
                self.sku_data = self.load_sheet_data(sku_sheet, 
                                                   self.sku_title_col_var.get(), 
                                                   self.sku_col_var.get())
                self.progress_var.set(50)
                self.root.update()
                
            # 加载成本数据
            if self.cost_sheet_var.get() and self.cost_sku_col_var.get() and self.cost_col_var.get():
                cost_sheet = self.workbook[self.cost_sheet_var.get()]
                self.cost_data = self.load_sheet_data(cost_sheet, 
                                                    self.cost_sku_col_var.get(), 
                                                    self.cost_col_var.get())
                self.progress_var.set(100)
                self.root.update()
                
            self.status_var.set(f"数据加载完成 - SKU数据: {len(self.sku_data)}条, 成本数据: {len(self.cost_data)}条")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载数据失败: {str(e)}")
            self.status_var.set("数据加载失败")
            
    def load_sheet_data(self, sheet, key_col, value_col):
        """加载工作表数据到字典"""
        data = {}
        key_col_num = openpyxl.utils.column_index_from_string(key_col)
        value_col_num = openpyxl.utils.column_index_from_string(value_col)
        
        for row in range(1, sheet.max_row + 1):
            key_cell = sheet.cell(row=row, column=key_col_num)
            value_cell = sheet.cell(row=row, column=value_col_num)
            
            if key_cell.value and str(key_cell.value).strip():
                # 清理键值，去除特殊字符
                clean_key = self.clean_text(str(key_cell.value))
                # 如果是SKU数据，也清理值
                if value_cell.value:
                    clean_value = self.clean_sku(str(value_cell.value))
                    data[clean_key] = clean_value
                else:
                    data[clean_key] = value_cell.value
                
        return data
        
    def clean_text(self, text):
        """清理文本，去除特殊字符"""
        if not text:
            return ""
            
        text = str(text)
        
        # 去除不可见字符和换行符
        text = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', text)
        
        # 去除多余空格
        text = ' '.join(text.split())
        
        # 去除首尾空格
        text = text.strip()
        
        return text
        
    def clean_sku(self, sku):
        """清理SKU，处理特殊字符和格式问题"""
        if not sku:
            return ""
            
        sku = str(sku)
        
        # 去除不可见字符
        sku = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', sku)
        
        # 去除首尾空格
        sku = sku.strip()
        
        # 统一空格处理（将多个空格替换为单个空格）
        sku = re.sub(r'\s+', ' ', sku)
        
        return sku
        
    def process_data(self):
        """处理数据"""
        try:
            if not self.sku_data or not self.cost_data:
                messagebox.showwarning("警告", "请先加载数据")
                return
                
            if not self.output_sheet_var.get():
                messagebox.showwarning("警告", "请选择输出工作表")
                return
                
            if not self.output_title_col_var.get() or not self.output_sku_col_var.get() or not self.output_cost_col_var.get():
                messagebox.showwarning("警告", "请选择输出工作表的标题列、SKU列和成本列")
                return
                
            self.status_var.set("正在处理数据...")
            self.progress_var.set(0)
            self.root.update()
            
            # 获取输出工作表
            output_sheet = self.workbook[self.output_sheet_var.get()]
            start_row = int(self.start_row_var.get())
            end_row = int(self.end_row_var.get())
            
            # 获取列号
            title_col_num = openpyxl.utils.column_index_from_string(self.output_title_col_var.get())
            sku_col_num = openpyxl.utils.column_index_from_string(self.output_sku_col_var.get())
            cost_col_num = openpyxl.utils.column_index_from_string(self.output_cost_col_var.get())
            
            total_rows = end_row - start_row + 1
            processed = 0
            found_sku_count = 0
            found_cost_count = 0
            
            # 处理每一行
            for row in range(start_row, end_row + 1):
                # 使用用户选择的列
                title_cell = output_sheet.cell(row=row, column=title_col_num)
                sku_cell = output_sheet.cell(row=row, column=sku_col_num)
                cost_cell = output_sheet.cell(row=row, column=cost_col_num)
                
                if title_cell.value and str(title_cell.value).strip():
                    # 清理标题文本
                    clean_title = self.clean_text(str(title_cell.value))
                    
                    # 查找SKU
                    if clean_title in self.sku_data:
                        sku_cell.value = self.sku_data[clean_title]
                        found_sku_count += 1
                        
                        # 根据SKU查找成本
                        sku_value = self.clean_sku(str(sku_cell.value))
                        if sku_value in self.cost_data:
                            cost_cell.value = self.cost_data[sku_value]
                            found_cost_count += 1
                        else:
                            cost_cell.value = "未找到成本"
                    else:
                        sku_cell.value = "未找到SKU"
                        cost_cell.value = "未找到成本"
                
                processed += 1
                progress = (processed / total_rows) * 100
                self.progress_var.set(progress)
                self.root.update()
                
            self.status_var.set(f"数据处理完成，共处理{processed}行，找到SKU: {found_sku_count}个，找到成本: {found_cost_count}个")
            
        except Exception as e:
            messagebox.showerror("错误", f"处理数据失败: {str(e)}")
            self.status_var.set("数据处理失败")
            
    def save_results(self):
        """保存结果"""
        try:
            if not self.workbook:
                messagebox.showwarning("警告", "没有数据可保存")
                return
                
            original_path = self.file_path_var.get()
            
            # 尝试保存到原文件
            try:
                self.workbook.save(original_path)
                messagebox.showinfo("成功", "结果已保存到原文件")
                self.status_var.set("结果已保存到原文件")
                
            except PermissionError:
                # 如果权限被拒绝，创建副本文件
                self.save_to_backup_file(original_path)
                
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {str(e)}")
            
    def save_to_backup_file(self, original_path):
        """保存到备份文件"""
        try:
            import os
            from datetime import datetime
            
            # 获取原文件信息
            file_dir = os.path.dirname(original_path)
            file_name = os.path.basename(original_path)
            name, ext = os.path.splitext(file_name)
            
            # 生成带时间戳的新文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"{name}_处理结果_{timestamp}{ext}"
            backup_path = os.path.join(file_dir, backup_name)
            
            # 保存到备份文件
            self.workbook.save(backup_path)
            
            # 显示成功信息
            result = messagebox.showinfo(
                "保存成功", 
                f"原文件正在被其他程序使用，已自动创建副本文件：\n\n{backup_name}\n\n文件位置：\n{file_dir}",
                type=messagebox.OK
            )
            
            self.status_var.set(f"结果已保存到副本文件: {backup_name}")
            
        except Exception as e:
            messagebox.showerror("错误", f"创建副本文件失败: {str(e)}")
            
    def test_save(self):
        """测试保存功能"""
        try:
            if not self.workbook:
                messagebox.showwarning("警告", "没有工作簿可测试")
                return
                
            # 获取输出工作表
            output_sheet = self.workbook[self.output_sheet_var.get()]
            
            # 在测试单元格写入数据
            test_cell = output_sheet.cell(row=1, column=1)
            test_cell.value = "测试数据 - " + str(datetime.now())
            
            original_path = self.file_path_var.get()
            
            # 尝试保存到原文件
            try:
                self.workbook.save(original_path)
                messagebox.showinfo("成功", "测试数据已保存到原文件，请检查Excel文件A1单元格")
                
            except PermissionError:
                # 如果权限被拒绝，创建副本文件
                self.save_test_to_backup_file(original_path)
            
        except Exception as e:
            messagebox.showerror("错误", f"测试保存失败: {str(e)}")
            
    def save_test_to_backup_file(self, original_path):
        """保存测试数据到备份文件"""
        try:
            import os
            from datetime import datetime
            
            # 获取原文件信息
            file_dir = os.path.dirname(original_path)
            file_name = os.path.basename(original_path)
            name, ext = os.path.splitext(file_name)
            
            # 生成带时间戳的测试文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            test_name = f"{name}_测试_{timestamp}{ext}"
            test_path = os.path.join(file_dir, test_name)
            
            # 保存到测试文件
            self.workbook.save(test_path)
            
            # 显示成功信息
            messagebox.showinfo(
                "测试保存成功", 
                f"原文件正在被其他程序使用，已创建测试副本文件：\n\n{test_name}\n\n文件位置：\n{file_dir}\n\n请检查A1单元格的测试数据"
            )
            
        except Exception as e:
            messagebox.showerror("错误", f"创建测试副本文件失败: {str(e)}")
            
    def show_help(self):
        """显示使用说明"""
        help_window = tk.Toplevel(self.root)
        help_window.title("📖 使用说明 - 资料准备与操作教程")
        help_window.geometry("1400x900")
        help_window.configure(bg='#2c3e50')
        
        # 创建主框架
        main_frame = tk.Frame(help_window, bg='#2c3e50')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 标题
        title_label = tk.Label(main_frame, text="📖 Excel数据处理工具使用说明", 
                              font=self.font_title, 
                              bg='#2c3e50', fg='#ecf0f1')
        title_label.pack(pady=(0, 20))
        
        # 创建标签页控件
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 资料准备标签页（文本 + 图片）
        prep_frame = tk.Frame(notebook, bg='#ecf0f1')
        notebook.add(prep_frame, text="📋 资料准备")
        self.create_help_with_images(
            prep_frame,
            self.get_data_preparation_content(),
            ["01.png", "02.png", "03.png"]
        )
        
        # 操作教程标签页
        tutorial_frame = tk.Frame(notebook, bg='#ecf0f1')
        notebook.add(tutorial_frame, text="🎯 操作教程")
        self.create_simple_help_content(tutorial_frame, self.get_operation_tutorial_content())
        
        # 常见问题标签页
        faq_frame = tk.Frame(notebook, bg='#ecf0f1')
        notebook.add(faq_frame, text="❓ 常见问题")
        self.create_simple_help_content(faq_frame, self.get_faq_content())
        
        # 关闭窗口
        def on_closing():
            help_window.destroy()
        help_window.protocol("WM_DELETE_WINDOW", on_closing)
        
    def create_simple_help_content(self, parent, content):
        """创建简单的帮助内容显示"""
        # 直接创建文本控件，带滚动条
        text_frame = tk.Frame(parent, bg='#ecf0f1')
        text_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 创建文本控件和滚动条
        content_text = tk.Text(text_frame, wrap=tk.WORD, 
                              font=('Microsoft YaHei', 11), 
                              bg='#ffffff', fg='#2c3e50',
                              relief='solid', bd=1,
                              padx=15, pady=15)
        
        scrollbar = tk.Scrollbar(text_frame, orient="vertical", command=content_text.yview)
        content_text.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        content_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 插入内容
        content_text.insert(tk.END, content)
        content_text.config(state=tk.DISABLED)
        
        
    def create_help_with_images(self, parent, content, image_files):
        """创建带图片的帮助内容显示
        image_files: List[str]，图片文件名，优先从 img/ 子目录查找，不存在则回退到项目根目录。
        """
        # 布局：上方文字说明，下方图片画廊（可滚动）
        outer = tk.Frame(parent, bg='#ecf0f1')
        outer.pack(fill='both', expand=True, padx=20, pady=20)

        # 文本区域
        text_frame = tk.Frame(outer, bg='#ecf0f1')
        text_frame.pack(fill='x', padx=0, pady=(0, 12))
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=('Microsoft YaHei', 11),
                              bg='#ffffff', fg='#2c3e50', relief='solid', bd=1, padx=15, pady=15, height=12)
        text_widget.pack(fill='x', expand=False)
        text_widget.insert(tk.END, content)
        text_widget.config(state=tk.DISABLED)

        # 图片滚动区域
        gallery_frame = tk.Frame(outer, bg='#ecf0f1')
        gallery_frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(gallery_frame, bg='#ecf0f1', highlightthickness=0)
        scrollbar = tk.Scrollbar(gallery_frame, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        inner = tk.Frame(canvas, bg='#ecf0f1')
        canvas.create_window((0, 0), window=inner, anchor='nw')

        # 记录PhotoImage防止被GC
        if not hasattr(self, '_help_images_refs'):
            self._help_images_refs = []

        def resolve_path(filename):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            candidate1 = os.path.join(base_dir, 'img', filename)
            candidate2 = os.path.join(base_dir, filename)
            if os.path.exists(candidate1):
                return candidate1
            if os.path.exists(candidate2):
                return candidate2
            return None

        def load_and_fit_image(path, max_width=1100):
            try:
                from PIL import Image, ImageTk
            except Exception:
                return None
            try:
                img = Image.open(path)
                w, h = img.size
                if w > max_width:
                    scale = max_width / float(w)
                    new_size = (int(w * scale), int(h * scale))
                    img = img.resize(new_size, Image.LANCZOS)
                return ImageTk.PhotoImage(img)
            except Exception:
                return None

        # 逐张图片加入
        for idx, name in enumerate(image_files, start=1):
            p = resolve_path(name)
            caption = f"步骤图 {idx}: {name}"

            cap_label = tk.Label(inner, text=caption, font=self.font_subtitle, bg='#ecf0f1', fg='#2c3e50')
            cap_label.pack(anchor='w', pady=(8, 4))

            if p:
                photo = load_and_fit_image(p)
                if photo:
                    img_label = tk.Label(inner, image=photo, bg='#ecf0f1')
                    img_label.pack(anchor='w', pady=(0, 12))
                    self._help_images_refs.append(photo)
                else:
                    tk.Label(inner, text=f"无法加载图片（需要Pillow库）：{p}", font=self.font_normal,
                             bg='#ecf0f1', fg='#7f8c8d').pack(anchor='w', pady=(0, 12))
            else:
                tk.Label(inner, text=f"未找到图片文件：{name}", font=self.font_normal,
                         bg='#ecf0f1', fg='#7f8c8d').pack(anchor='w', pady=(0, 12))

        # 更新滚动区域大小
        def on_configure(event=None):
            inner.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox('all'))

        inner.bind('<Configure>', on_configure)
        on_configure()

    def get_data_preparation_content(self):
        """获取资料准备内容"""
        return """📊 Excel文件结构要求：

1. 工作表命名建议：
   • SKU数据源：Sheet1  建议从Upseller产品列表导出对应店铺Excel工作表
   • 成本数据源：Sheet2  建议从Upseller库存清单导出对应店铺Excel工作表  
   • 输出工作表：Order details 建议Tk后台导出Excel工作表

2. SKU数据源结构 (Sheet1)：
   • A列：产品ID (可选)
   • B列：产品标题 (必填) - 用于搜索SKU
   • C列：SKU数据 (必填) - 查找结果
   • D列：变种ID (可选)

3. 成本数据源结构 (Sheet2)：
   • A列：SKU标识 (必填) - 用于匹配
   • B列：平均成本 (必填) - 查找结果
   • C列：标题 (可选)
   • D列：仓库 (可选)
   • E列：货架位 (可选)

4. 输出工作表结构 (Order details)：
   • A列：Product name (产品标题) - 输入数据
   • B列：SKU (输出位置) - 处理结果
   • C列：Quantity (数量)
   • D列：成本 (输出位置) - 处理结果
   • E列：总成本

📝 数据质量要求：
• 确保产品标题完整，无空值
• SKU格式统一，避免特殊字符
• 成本数据为数值格式
• 建议数据量不超过10000行"""
        
    def get_operation_tutorial_content(self):
        """获取操作教程内容"""
        return """🚀 详细操作步骤：

第一步：启动程序
1. 双击运行 excel_processor.py 或启动程序.bat
2. 等待程序加载完成

第二步：选择Excel文件
1. 点击"📁 浏览文件"按钮
2. 选择包含数据的Excel文件
3. 程序自动加载所有工作表

第三步：自动配置 (推荐)
1. 点击"🎯 自动配置"按钮
2. 程序自动设置最佳配置
3. 查看配置说明确认无误

第四步：数据检查 (重要)
1. 点击"🔍 数据检查"按钮
2. 查看匹配率统计
3. 检查不匹配的SKU列表
4. 如匹配率低于80%，建议手动检查数据

第五步：加载数据
1. 点击"📥 加载数据"按钮
2. 等待数据加载完成
3. 查看状态栏的加载结果

第六步：开始处理
1. 点击"⚡ 开始处理"按钮
2. 观察进度条和状态信息
3. 等待处理完成

第七步：保存结果
1. 点击"💾 保存结果"按钮
2. 如原文件被占用，会自动创建副本
3. 查看保存成功提示

🔧 高级功能：
• 调试信息：查看详细配置和数据信息
• 测试保存：验证文件保存功能
• 手动配置：自定义列设置"""
        
    def get_faq_content(self):
        """获取常见问题内容"""
        return """❓ 常见问题解答：

Q1: 程序提示"未找到SKU"或"未找到成本"？
A1: 可能原因：
    • SKU格式不一致（特殊字符、空格、大小写）
    • 数据源配置错误
    • 数据清理问题
    解决方案：使用"数据检查"功能分析匹配情况

Q2: 保存时提示权限错误？
A2: 程序会自动创建副本文件，无需担心数据丢失
    副本文件命名：原文件名_处理结果_时间戳.xlsx

Q3: 处理速度很慢？
A3: 优化建议：
    • 减少处理行数范围
    • 关闭其他占用内存的程序
    • 分批处理大量数据

Q4: 自动配置不准确？
A4: 可以手动调整：
    • 重新选择工作表
    • 修改列设置
    • 使用"调试信息"查看配置

Q5: 数据匹配率低？
A5: 检查要点：
    • SKU格式是否统一
    • 是否有特殊字符
    • 数据源是否正确
    • 使用"数据检查"详细分析

Q6: 程序无法启动？
A6: 检查项目：
    • Python版本 (建议3.7+)
    • 依赖库是否正确安装
    • Excel文件格式是否正确

📞 技术支持：
如遇到其他问题，请检查：
• 错误提示信息
• 数据格式是否正确
• 文件权限是否正常"""
            
    def show_debug_info(self):
        """显示调试信息"""
        try:
            debug_info = []
            debug_info.append("=== 调试信息 ===")
            debug_info.append(f"文件路径: {self.file_path_var.get()}")
            debug_info.append(f"工作簿状态: {'已加载' if self.workbook else '未加载'}")
            
            if self.workbook:
                debug_info.append(f"工作表数量: {len(self.workbook.sheetnames)}")
                debug_info.append(f"工作表列表: {', '.join(self.workbook.sheetnames)}")
            
            debug_info.append(f"SKU数据源: {self.sku_sheet_var.get()} - {self.sku_title_col_var.get()} -> {self.sku_col_var.get()}")
            debug_info.append(f"成本数据源: {self.cost_sheet_var.get()} - {self.cost_sku_col_var.get()} -> {self.cost_col_var.get()}")
            debug_info.append(f"输出设置: {self.output_sheet_var.get()} - 标题:{self.output_title_col_var.get()}, SKU:{self.output_sku_col_var.get()}, 成本:{self.output_cost_col_var.get()}")
            
            debug_info.append(f"SKU数据条数: {len(self.sku_data) if self.sku_data else 0}")
            debug_info.append(f"成本数据条数: {len(self.cost_data) if self.cost_data else 0}")
            
            if self.sku_data:
                debug_info.append("SKU数据示例:")
                for i, (key, value) in enumerate(list(self.sku_data.items())[:3]):
                    debug_info.append(f"  {key} -> {value}")
                if len(self.sku_data) > 3:
                    debug_info.append(f"  ... 还有{len(self.sku_data)-3}条数据")
            
            if self.cost_data:
                debug_info.append("成本数据示例:")
                for i, (key, value) in enumerate(list(self.cost_data.items())[:3]):
                    debug_info.append(f"  {key} -> {value}")
                if len(self.cost_data) > 3:
                    debug_info.append(f"  ... 还有{len(self.cost_data)-3}条数据")
            
            # 显示调试信息窗口
            debug_window = tk.Toplevel(self.root)
            debug_window.title("调试信息")
            debug_window.geometry("600x500")
            
            text_widget = tk.Text(debug_window, wrap=tk.WORD, font=("微软雅黑", 9))
            scrollbar = tk.Scrollbar(debug_window, orient=tk.VERTICAL, command=text_widget.yview)
            text_widget.configure(yscrollcommand=scrollbar.set)
            
            text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            text_widget.insert(tk.END, '\n'.join(debug_info))
            text_widget.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("错误", f"显示调试信息失败: {str(e)}")
            
    def auto_config(self):
        """自动配置列设置"""
        try:
            if not self.workbook:
                messagebox.showwarning("警告", "请先选择Excel文件")
                return
                
            # 自动设置SKU数据源
            if 'Sheet1' in self.workbook.sheetnames:
                self.sku_sheet_var.set('Sheet1')
                self.sku_title_col_var.set('B')  # 产品标题列
                self.sku_col_var.set('C')        # SKU列
                
            # 自动设置成本数据源
            if 'Sheet2' in self.workbook.sheetnames:
                self.cost_sheet_var.set('Sheet2')
                self.cost_sku_col_var.set('A')   # SKU列
                self.cost_col_var.set('B')       # 成本列
                
            # 自动设置输出
            if 'Order details' in self.workbook.sheetnames:
                self.output_sheet_var.set('Order details')
                self.output_title_col_var.set('A')  # 产品标题列
                self.output_sku_col_var.set('B')    # SKU列
                self.output_cost_col_var.set('D')   # 成本列
                
            messagebox.showinfo("成功", "自动配置完成！\n\n配置说明：\n- SKU数据源：Sheet1 (B列标题 -> C列SKU)\n- 成本数据源：Sheet2 (A列SKU -> B列成本)\n- 输出：Order details (A列标题 -> B列SKU -> D列成本)")
            
        except Exception as e:
            messagebox.showerror("错误", f"自动配置失败: {str(e)}")
            
    def check_data_consistency(self):
        """检查数据一致性"""
        try:
            if not self.workbook:
                messagebox.showwarning("警告", "请先选择Excel文件")
                return
                
            if not self.sku_sheet_var.get() or not self.cost_sheet_var.get():
                messagebox.showwarning("警告", "请先配置SKU数据源和成本数据源")
                return
                
            self.status_var.set("正在检查数据一致性...")
            self.progress_var.set(0)
            self.root.update()
            
            # 加载SKU数据
            sku_sheet = self.workbook[self.sku_sheet_var.get()]
            sku_data = {}
            sku_title_col_num = openpyxl.utils.column_index_from_string(self.sku_title_col_var.get())
            sku_col_num = openpyxl.utils.column_index_from_string(self.sku_col_var.get())
            
            for row in range(2, sku_sheet.max_row + 1):
                title_cell = sku_sheet.cell(row=row, column=sku_title_col_num)
                sku_cell = sku_sheet.cell(row=row, column=sku_col_num)
                
                if title_cell.value and sku_cell.value:
                    clean_title = self.clean_text(str(title_cell.value))
                    clean_sku = self.clean_sku(str(sku_cell.value))
                    sku_data[clean_title] = clean_sku
            
            self.progress_var.set(50)
            self.root.update()
            
            # 加载成本数据
            cost_sheet = self.workbook[self.cost_sheet_var.get()]
            cost_data = {}
            cost_sku_col_num = openpyxl.utils.column_index_from_string(self.cost_sku_col_var.get())
            cost_col_num = openpyxl.utils.column_index_from_string(self.cost_col_var.get())
            
            for row in range(2, cost_sheet.max_row + 1):
                sku_cell = cost_sheet.cell(row=row, column=cost_sku_col_num)
                cost_cell = cost_sheet.cell(row=row, column=cost_col_num)
                
                if sku_cell.value and cost_cell.value:
                    clean_sku = self.clean_sku(str(sku_cell.value))
                    cost_data[clean_sku] = str(cost_cell.value)
            
            self.progress_var.set(100)
            self.root.update()
            
            # 分析数据一致性
            sku_values = set(sku_data.values())
            cost_keys = set(cost_data.keys())
            
            matched_skus = sku_values.intersection(cost_keys)
            unmatched_skus = sku_values - cost_keys
            unused_cost_skus = cost_keys - sku_values
            
            # 显示检查结果窗口
            self.show_consistency_report(sku_data, cost_data, matched_skus, unmatched_skus, unused_cost_skus)
            
            self.status_var.set(f"数据检查完成 - 匹配: {len(matched_skus)}, 不匹配: {len(unmatched_skus)}")
            
        except Exception as e:
            messagebox.showerror("错误", f"数据检查失败: {str(e)}")
            self.status_var.set("数据检查失败")
            
    def show_consistency_report(self, sku_data, cost_data, matched_skus, unmatched_skus, unused_cost_skus):
        """显示数据一致性报告"""
        report_window = tk.Toplevel(self.root)
        report_window.title("数据一致性检查报告")
        report_window.geometry("800x600")
        
        # 创建笔记本控件（标签页）
        notebook = ttk.Notebook(report_window)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # 总览标签页
        overview_frame = ttk.Frame(notebook)
        notebook.add(overview_frame, text="总览")
        
        overview_text = tk.Text(overview_frame, wrap=tk.WORD, font=("微软雅黑", 9))
        overview_scrollbar = tk.Scrollbar(overview_frame, orient=tk.VERTICAL, command=overview_text.yview)
        overview_text.configure(yscrollcommand=overview_scrollbar.set)
        
        overview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        overview_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        overview_info = f"""数据一致性检查报告
{'='*50}

SKU数据源统计:
- 总数据条数: {len(sku_data)}
- 唯一SKU数量: {len(set(sku_data.values()))}

成本数据源统计:
- 总数据条数: {len(cost_data)}
- 唯一SKU数量: {len(cost_data)}

匹配情况:
- 匹配的SKU: {len(matched_skus)} 个
- 不匹配的SKU: {len(unmatched_skus)} 个
- 未使用的成本SKU: {len(unused_cost_skus)} 个

匹配率: {len(matched_skus) / len(set(sku_data.values())) * 100:.1f}%

建议:
1. 如果匹配率低于80%，建议检查SKU格式是否一致
2. 检查是否有特殊字符、空格、大小写等问题
3. 手动修正不匹配的SKU数据
"""
        
        overview_text.insert(tk.END, overview_info)
        overview_text.config(state=tk.DISABLED)
        
        # 不匹配SKU标签页
        if unmatched_skus:
            unmatched_frame = ttk.Frame(notebook)
            notebook.add(unmatched_frame, text=f"不匹配SKU ({len(unmatched_skus)})")
            
            unmatched_text = tk.Text(unmatched_frame, wrap=tk.WORD, font=("微软雅黑", 9))
            unmatched_scrollbar = tk.Scrollbar(unmatched_frame, orient=tk.VERTICAL, command=unmatched_text.yview)
            unmatched_text.configure(yscrollcommand=unmatched_scrollbar.set)
            
            unmatched_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            unmatched_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            unmatched_info = "不匹配的SKU列表:\n" + "="*50 + "\n\n"
            for sku in sorted(unmatched_skus):
                # 找到对应的标题
                titles = [title for title, s in sku_data.items() if s == sku]
                unmatched_info += f"SKU: {sku}\n"
                unmatched_info += f"标题: {titles[0] if titles else '未知'}\n"
                unmatched_info += f"问题: 在成本数据源中未找到此SKU\n"
                unmatched_info += "-" * 30 + "\n"
            
            unmatched_text.insert(tk.END, unmatched_info)
            unmatched_text.config(state=tk.DISABLED)
        
        # 未使用成本SKU标签页
        if unused_cost_skus:
            unused_frame = ttk.Frame(notebook)
            notebook.add(unused_frame, text=f"未使用成本SKU ({len(unused_cost_skus)})")
            
            unused_text = tk.Text(unused_frame, wrap=tk.WORD, font=("微软雅黑", 9))
            unused_scrollbar = tk.Scrollbar(unused_frame, orient=tk.VERTICAL, command=unused_text.yview)
            unused_text.configure(yscrollcommand=unused_scrollbar.set)
            
            unused_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            unused_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            unused_info = "未使用的成本SKU列表:\n" + "="*50 + "\n\n"
            for sku in sorted(list(unused_cost_skus)[:50]):  # 只显示前50个
                cost_value = cost_data.get(sku, '未知')
                unused_info += f"SKU: {sku}\n"
                unused_info += f"成本: {cost_value}\n"
                unused_info += f"问题: 在SKU数据源中未找到此SKU\n"
                unused_info += "-" * 30 + "\n"
            
            if len(unused_cost_skus) > 50:
                unused_info += f"\n... 还有 {len(unused_cost_skus) - 50} 个未显示的SKU"
            
            unused_text.insert(tk.END, unused_info)
            unused_text.config(state=tk.DISABLED)
        
        # 匹配示例标签页
        if matched_skus:
            matched_frame = ttk.Frame(notebook)
            notebook.add(matched_frame, text=f"匹配示例 ({len(matched_skus)})")
            
            matched_text = tk.Text(matched_frame, wrap=tk.WORD, font=("微软雅黑", 9))
            matched_scrollbar = tk.Scrollbar(matched_frame, orient=tk.VERTICAL, command=matched_text.yview)
            matched_text.configure(yscrollcommand=matched_scrollbar.set)
            
            matched_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            matched_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            matched_info = "成功匹配的SKU示例:\n" + "="*50 + "\n\n"
            for sku in sorted(list(matched_skus)[:20]):  # 只显示前20个
                titles = [title for title, s in sku_data.items() if s == sku]
                cost_value = cost_data.get(sku, '未知')
                matched_info += f"SKU: {sku}\n"
                matched_info += f"标题: {titles[0] if titles else '未知'}\n"
                matched_info += f"成本: {cost_value}\n"
                matched_info += "状态: ✓ 匹配成功\n"
                matched_info += "-" * 30 + "\n"
            
            if len(matched_skus) > 20:
                matched_info += f"\n... 还有 {len(matched_skus) - 20} 个匹配的SKU"
            
            matched_text.insert(tk.END, matched_info)
            matched_text.config(state=tk.DISABLED)

def main():
    root = tk.Tk()
    app = ExcelProcessor(root)
    root.mainloop()

if __name__ == "__main__":
    main()
